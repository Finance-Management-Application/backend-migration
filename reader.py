import logging

from openpyxl import load_workbook

from .constants import Workbook 
from .constants import WORKBOOK_SHEET_PYDANTIC_MODEL_MAPPING as mapping

def read(sheet:Workbook):
    wb = load_workbook(filename=Workbook.LOCATION, read_only=True, data_only=True)
    ws = wb[sheet]
    
    columns = [cell.value for cell in next(ws.rows)]
    keys = ["_".join(column.lower().split()) for column in columns]

    data = []
    for row in ws.iter_rows(min_row=2):
        record = {key: cell.value for key, cell in zip(keys, row)}
        model = mapping.get(sheet)
        if model:
            data.append(model(**record).model_dump())
    
    return data