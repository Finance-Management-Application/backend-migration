from typing import List, Tuple, Dict, Any
from pathlib import Path
from dataclasses import dataclass
from concurrent.futures import ThreadPoolExecutor
from functools import partial

from pydantic import BaseModel, ValidationError
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from logging import DEBUG
from logconfig import get_logger

logger = get_logger(__name__, f'{__file__[:-3]}.log', DEBUG)

@dataclass
class ExcelReadResult:
    """Container for Excel reading results"""
    data: List[Dict[str, Any]]
    success_rate: float
    success_count: int
    failure_count: int

class ExcelReaderError(Exception):
    """Base exception class for Excel reader errors"""
    pass

class WorkbookError(ExcelReaderError):
    """Exception for workbook-related errors"""
    pass

class SheetError(ExcelReaderError):
    """Exception for worksheet-related errors"""
    pass

class SchemaError(ExcelReaderError):
    """Exception for schema-related errors"""
    pass

def validate_inputs(workbook: Path, max_col: int, model: BaseModel) -> None:
    """Validate input parameters"""
    if not isinstance(workbook, Path):
        raise TypeError(f'workbook must be a Path object, got {type(workbook)}')
    if not isinstance(max_col, int) or max_col <= 0:
        raise TypeError(f'max_col must be a positive integer, got {max_col}')
    if not issubclass(type(model), type(BaseModel)):
        raise TypeError(f'model must be a Pydantic BaseModel class, got {type(model)}')

def process_headers(ws, max_col: int, workbook: Path) -> List[str]:
    """Process worksheet headers and return normalized column names"""
    try:
        first_row = next(ws.iter_rows(max_col=max_col))
        columns = [cell.value for cell in first_row]
        
        if not all(columns):
            raise SchemaError(f'{workbook.name}: Found empty header cells')
            
        return ["_".join(str(column).lower().split()) for column in columns]
    except StopIteration:
        raise SheetError(f'{workbook.name}: Sheet is empty')
    except AttributeError:
        raise SchemaError(f'{workbook.name}: Invalid max_col value: {max_col}')

def process_row(row: tuple, keys: List[str], model: BaseModel, row_num: int) -> Dict[str, Any]:
    """Process a single row and return validated data"""
    record = {}
    for key, cell in zip(keys, row):
        record[key] = cell.value
        
    try:
        return model(**record).model_dump()
    except ValidationError as e:
        logger.debug(f'Validation error in row {row_num}: {e}')
        raise

def read(workbook: Path, max_col: int, model: BaseModel) -> ExcelReadResult:
    """
    Read and validate data from a single Excel workbook.
    
    Args:
        workbook: Path to Excel file
        max_col: Maximum number of columns to read
        model: Pydantic model for data validation
    
    Returns:
        ExcelReadResult containing processed data and statistics
    """
    validate_inputs(workbook, max_col, model)
    
    try:
        wb = load_workbook(filename=workbook, read_only=True, data_only=True)
    except FileNotFoundError:
        raise WorkbookError(f'{workbook.name}: File not found')
    except InvalidFileException:
        raise WorkbookError(f'{workbook.name}: Invalid Excel format')
    except Exception as e:
        raise WorkbookError(f'{workbook.name}: Error opening file - {str(e)}')

    ws = wb.active
    keys = process_headers(ws, max_col, workbook)
    
    data, success, failure = [], 0, 0
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_col=max_col), start=2):
        try:
            validated_row = process_row(row, keys, model, row_num)
            data.append(validated_row)
            success += 1
        except (ValidationError, IndexError):
            failure += 1
            continue
    
    total = success + failure
    success_rate = (success / total * 100) if total > 0 else 0
    
    return ExcelReadResult(
        data=data,
        success_rate=round(success_rate, 2),
        success_count=success,
        failure_count=failure
    )

def read_multiple(workbooks: List[Path], max_col: int, model: BaseModel, 
                 max_workers: int = None) -> Tuple[List[Dict[str, Any]], List[Dict[str, float]]]:
    """
    Read multiple Excel workbooks in parallel.
    
    Args:
        workbooks: List of paths to Excel files
        max_col: Maximum number of columns to read
        model: Pydantic model for data validation
        max_workers: Maximum number of parallel workers
    
    Returns:
        Tuple of (combined data, success rates per workbook)
    """
    if not isinstance(workbooks, list):
        raise TypeError(f'workbooks must be a list, got {type(workbooks)}')
    
    partial_read = partial(read, max_col=max_col, model=model)
    
    all_data = []
    stats = []
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(partial_read, wb) for wb in workbooks]
        
        for workbook, future in zip(workbooks, futures):
            try:
                result = future.result()
                all_data.extend(result.data)
                stats.append({workbook.name: result.success_rate})
            except ExcelReaderError as e:
                logger.error(f'Error processing {workbook.name}: {str(e)}')
                stats.append({workbook.name: 0})
            except Exception as e:
                logger.error(f'Unexpected error processing {workbook.name}: {str(e)}')
                stats.append({workbook.name: 0})
    
    return all_data, stats