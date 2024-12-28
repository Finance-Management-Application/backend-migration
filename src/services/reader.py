from typing import List, Tuple, Union, Generator
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

from pydantic import BaseModel
from pydantic import ValidationError
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from logging import DEBUG
from logconfig import get_logger # type: ignore


logger = get_logger(__name__, f'{__file__[:-3]}.log', DEBUG)


def read(workbook: Path, max_col: int, model: BaseModel) -> Tuple[List[dict], dict]:

    data, failure, success = [], 0, 0

    try:
        if not isinstance(workbook, Path):
            logger.error(f'FunctionParameterError: {workbook} must be a {Path} type Object')
            raise TypeError
        
        if not isinstance(max_col, int):
            logger.error(f'FunctionParameterError: max_col must be a {int} type Object')
            raise TypeError
        elif max_col <=0:
            logger.exception(f'FunctionParameterError: max_col must be a positive integer, got {max_col}')
            raise ValueError
        
        if not issubclass(model, BaseModel):
            logger.error(f'FunctionParameterError: {model} must be a subclass of {BaseModel} type Object')
            raise
        

        try:
            wb = load_workbook(filename=workbook, read_only=True, data_only=True)
        except FileNotFoundError as e:
            logger.exception(f'WorkbookError:{workbook.name}: File not found in {workbook} Path')
            raise
        except InvalidFileException as e:
            logger.exception(f'WorkbookError:{workbook.name}: Workbook is non-ooxml(not xlsx format in this case) file')
            raise
        except Exception as e:
            logger.exception(f'WorkbookError:{workbook.name}: Unknown Error in opening')
            raise
        
        ws = wb.active

        try:
            column_headers = [cell.value for cell in next(ws.iter_rows(max_col=max_col))]
            if not all(column_headers):
                logger.exception(f'SchemaError:{workbook.name}: Found empty header cells')
                raise # I should write a custom exception class for this
        except StopIteration as e:
            logger.exception(f'SheetError:{workbook.name}: {str(ws)} is Empty')
            raise
        else:
            keys = ["_".join(column.lower().split()) for column in column_headers]

        
        
        for n, row in enumerate(ws.iter_rows(min_row=2, max_col=max_col), start=2):
            record = { key: cell.value for key, cell in zip(keys, row) }
            try:
                data.append(model(**record).model_dump())
                success+=1
            except ValidationError as e:
                logger.debug(f'DataError:{workbook.name}: Row={n} \n {e}')
                failure+=1
                continue

    except (TypeError, ValueError, StopIteration, FileNotFoundError, 
            ValidationError, InvalidFileException) as e:
        logger.info(f'WorkbookError:{workbook.name}: Cannot be processed.')
    finally:
        return (data, {workbook.name: round(success/(success+failure)*100)}) if success>0 else (data, 0)

def reads(workbooks: Union[List[Path], Generator], max_col: int, model: BaseModel) -> Tuple[List[dict], List[dict]]:
    
    data, stats  = [], []
    try:
        # isinstance(workbooks, types.GeneratorType)
        if not (isinstance(workbooks, list) or hasattr(workbooks, '__iter__')):
            raise TypeError
    except TypeError:
        logger.error(f'FunctionParameterError: {workbooks} must be a list or generator type Object')
    else:
        with ThreadPoolExecutor() as executor:
            futures = [executor.submit(read, workbook, max_col, model) for workbook in workbooks]
            for future in as_completed(futures):
                wb_data, stat = future.result()
                data.extend(wb_data)
                stats.append(stat)

            # NEED TO SPEND MORE TIME - ITS NOT WORKING
            # results = executor.map(read, workbooks, [max_col] * len(workbooks), [model] * len(workbooks))
            # for wb_data, stat in results:
            #     data.extend(wb_data)
            #     stats.append(stat)

        # for workbook in workbooks:
        #     wb_data, stat = read(workbook, max_col, model)
        #     data.extend(wb_data)
        #     stats.append(stat)       
    finally:
        return data, stats