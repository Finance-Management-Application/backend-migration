from typing import List, Union, Optional
from pathlib import Path

from pydantic import BaseModel
from pydantic import ValidationError
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from logging import DEBUG
from logconfig import get_logger

logger = get_logger(__name__, __file__, DEBUG)

def read(workbook: Path, max_col: int, model: BaseModel) -> List[dict]:  
    """
    Read and validate data from a single Excel workbook.
    
    Args:
        workbook: Path to Excel workbook
        max_col: Number of columns to read
        model: Pydantic model for validation
        
    Returns:
        List of validated data dictionaries
        
    Raises:
        TypeError: If input parameters are of wrong type
        FileNotFoundError: If Excel file doesn't exist
        PermissionError: If no permission to read file
        InvalidFileException: If file is not a valid Excel file
        ValueError: If Excel file is empty or has invalid structure
        AttributeError: If column headers contain None values
        KeyError: If row data doesn't match column headers
        ValidationError: If row data fails Pydantic validation
    """    
    # Validate input parameters
    if not isinstance(workbook, Path):
        raise TypeError(f"workbook must be Path object, got {type(workbook)}")
    if not isinstance(max_col, int):
        raise TypeError(f"max_col must be integer, got {type(max_col)}")
    if not isinstance(model, type) or not issubclass(model, BaseModel):
        raise TypeError(f"model must be Pydantic BaseModel class")
        
    try:
        # Check file existence
        if not workbook.exists():
            raise FileNotFoundError(f"Excel file not found: {workbook}")
            
        # Open workbook - may raise InvalidFileException if format invalid
        try:
            wb = load_workbook(filename=workbook, read_only=True, data_only=True)
        except InvalidFileException as e:
            logger.error(f"Invalid Excel file {workbook}: {str(e)}")
            raise
        except PermissionError as e:
            logger.error(f"Permission denied accessing {workbook}: {str(e)}")
            raise
        except Exception as e:
            logger.error(f"Error opening workbook {workbook}: {str(e)}")
            raise ValueError(f"Could not open Excel file: {str(e)}")

        ws = wb.active
        
        # Get headers - handle empty files
        try:
            first_row = next(ws.iter_rows(max_col=max_col))
            if not first_row:
                raise ValueError("Excel file is empty")
        except StopIteration:
            raise ValueError("Excel file is empty")
            
        # Process headers - handle None values
        try:
            columns = [cell.value for cell in first_row]
            if None in columns:
                raise ValueError(f"Found empty column header(s) in {workbook}")
            keys = ["_".join(str(column).lower().split()) for column in columns]
        except AttributeError as e:
            logger.error(f"Invalid column header in {workbook}: {str(e)}")
            raise

        data = []
        # Process data rows
        for n, row in enumerate(ws.iter_rows(min_row=2, max_col=max_col), start=2):
            try:
                # Create record dict safely
                record = {}
                for key, cell in zip(keys, row):
                    if not key:  # Handle case where there are more row cells than headers
                        raise KeyError(f"Row {n} has more columns than headers")
                    record[key] = cell.value if cell else None
                    
                # Validate with Pydantic model
                try:
                    validated_data = model(**record).model_dump()
                    data.append(validated_data)
                except ValidationError as e:
                    logger.debug(f'EXCEL FILE: {workbook.stem} ROW: {n} \n Validation Error: {e}')
                    continue  # Skip invalid rows but continue processing
                    
            except KeyError as e:
                logger.error(f"Row structure error in {workbook} row {n}: {str(e)}")
                continue
            except Exception as e:
                logger.error(f"Unexpected error processing row {n} in {workbook}: {str(e)}")
                continue

        return data
        
    finally:
        # Ensure workbook is closed even if exceptions occur
        try:
            wb.close()
        except:
            pass

def reads(workbooks: List[Path], columns: int, model: BaseModel) -> List[dict]:
    """
    Read and validate data from multiple Excel workbooks.
    
    Args:
        workbooks: List of paths to Excel workbooks
        columns: Number of columns to read
        model: Pydantic model for validation
        
    Returns:
        List of validated data dictionaries
        
    Raises:
        TypeError: If workbooks is not a list
        ValueError: If workbooks list is empty
    """
    # Validate input
    if not isinstance(workbooks, list):
        raise TypeError(f"workbooks must be a list, got {type(workbooks)}")
    if not workbooks:
        raise ValueError("workbooks list is empty")
        
    data = []
    for workbook in workbooks:
        try:
            workbook_data = read(workbook, columns, model)
            data.extend(workbook_data)
        except (FileNotFoundError, InvalidFileException, 
                PermissionError, ValueError) as e:
            logger.error(f"Error processing {workbook}: {str(e)}")
            continue  # Skip problematic files but continue with others
        except Exception as e:
            logger.error(f"Unexpected error with {workbook}: {str(e)}")
            continue
            
    return data